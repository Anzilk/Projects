from odoo import models, api
import base64
from pathlib import Path
import openpyxl
from openpyxl.styles import Color, Fill, Font, Alignment
import io


class AttachmentSpreadsheet(models.Model):
    _inherit = 'ir.attachment'

    def get_doc_file_data(self, *args):
        attachment = self.env['ir.attachment'].search([('id', '=', args[0])])
        bin_data = base64.b64decode(attachment.datas)
        # wb = xlrd.open_workbook(file_contents=bin_data)
        data = io.BytesIO(bin_data)
        wb = openpyxl.load_workbook(data)
        # sheet = wb.sheets()[0]
        # for s in wb.sheets():
        sheet = wb.worksheets[0]
        values = []
        bold_values = []
        italic_values = []
        underline_values = []
        font_size_values = []
        alignment_values = []
        vertical_values = []
        row_count = sheet.max_row
        column_count = sheet.max_column
        for row in range(row_count):
            col_value = []
            bold_value = []
            italics_value = []
            underline_value = []
            font_size_value = []
            alignment_value = []
            vertical_value = []
            for col in range(column_count):
                cell_value = sheet.cell(row+1, col+1)
                value = sheet.cell(row+1, col+1).value
                # print(cell_value.font,"a1 font style ")
                bold = cell_value.font.bold
                italics = cell_value.font.italic
                underline = cell_value.font.underline
                font_size = cell_value.font.size
                alignment = cell_value.alignment.horizontal
                vertical_align = cell_value.alignment.vertical
                # font_color = cell_value.font.color
                # if font_color:
                #     print(font_color.rgb,"colour new")
                # print(font_color,"colour")
                # try:
                #     value = str(int(value))
                # except:
                #     pass
                col_value.append(value)
                bold_value.append(bold)
                italics_value.append(italics)
                underline_value.append(underline)
                font_size_value.append(font_size)
                alignment_value.append(alignment)
                vertical_value.append(vertical_align)
            values.append(col_value)
            bold_values.append(bold_value)
            italic_values.append(italics_value)
            underline_values.append(underline_value)
            font_size_values.append(font_size_value)
            alignment_values.append(alignment_value)
            vertical_values.append(vertical_value)
        #     print(values,"values")
        return values, bold_values, italic_values, underline_values,\
               font_size_values, alignment_values, vertical_values

    @api.model
    def add_data_spreadsheet(self, f_value,
                             doc_id, r_value=0, c_value=0, fontsize=13,
                             boldvalue='False', italicvalue='False',
                             underline_value='none', align_value="left",
                             vertical_align_value="bottom"):
        row_value = int(r_value)
        col_value = int(c_value)
        # print(fontsize, "fontsize")
        if fontsize == '':
            fontsize=13;
        # print(fontsize,"fontsizezz")
        # fontsize = int(fontsize)
        attachment = self.env['ir.attachment'].browse(doc_id)
        bin_data = base64.b64decode(attachment.datas)
        # f = open(file_path, 'wb')
        data = io.BytesIO(bin_data)
        wb = openpyxl.load_workbook(data)
        sheet = wb.active
        sheet.cell(row_value, col_value).value = f_value
        cell_value = sheet.cell(row_value, col_value)
        cell_value.font = Font(size=fontsize, bold=boldvalue,
                               italic=italicvalue, underline=underline_value)
        cell_value.alignment = Alignment(horizontal=align_value,
                                         vertical=vertical_align_value)
        wb.save(Path(Path.home(), 'sample.xlsx'))

        stp_file = Path(Path.home(), 'sample.xlsx')
        stp_object_file = open(stp_file, 'rb')
        stp_object_file_encode = base64.b64encode(stp_object_file.read())
        stp_object_file_encode = stp_object_file_encode.decode('utf-8')
        attachment.write({'datas': stp_object_file_encode})

        return True
